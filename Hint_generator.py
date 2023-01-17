import json
import random
import openpyxl
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill

# find all json files (spoiler log) in the current directory
json_files = glob.glob('*.json')

# load data from the first json file
if json_files:
    with open(json_files[0], 'r') as f:
        data = json.load(f)
else:
    print("No json files found in current directory")

# Create a new xlsx file
workbook = openpyxl.Workbook()

# Select the active sheet
worksheet = workbook.active

# Write headers
worksheet.cell(row=2, column=2, value="WotH")
worksheet.cell(row=7, column=2, value="Barren")
worksheet.cell(row=11, column=2, value="Always")
worksheet.cell(row=17, column=2, value="Sometimes")

# Set the width of column B
worksheet.column_dimensions["B"].width = 69

# Add borders to cells B2 to B5
for i in range(2, 6):
    cell = worksheet.cell(row=i, column=2)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

# Add borders to cells B7 to B9
for i in range(7, 10):
    cell = worksheet.cell(row=i, column=2)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
                        
# Add borders to cells B11 to B15
for i in range(11, 16):
    cell = worksheet.cell(row=i, column=2)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

# Add borders to cells B17 to B20
for i in range(17, 21):
    cell = worksheet.cell(row=i, column=2)
    cell.border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
                        
# Header
worksheet["B2"].font = openpyxl.styles.Font(bold=True)
worksheet["B7"].font = openpyxl.styles.Font(bold=True)
worksheet["B11"].font = openpyxl.styles.Font(bold=True)
worksheet["B17"].font = openpyxl.styles.Font(bold=True)
worksheet["B2"].fill = PatternFill("solid", start_color="EBDD53")
worksheet["B7"].fill = PatternFill("solid", start_color="FF7979")
worksheet["B11"].fill = PatternFill("solid", start_color="C8E191")
worksheet["B17"].fill = PatternFill("solid", start_color="F9B883")

# Set the important zones to italic
for i in range(3, 5):
    worksheet.cell(row=i, column=2).font = openpyxl.styles.Font(italic=True)
for i in range(8, 9):
    worksheet.cell(row=i, column=2).font = openpyxl.styles.Font(italic=True)
for i in range(12, 15):
    worksheet.cell(row=i, column=2).font = openpyxl.styles.Font(italic=True)
for i in range(18, 20):
    worksheet.cell(row=i, column=2).font = openpyxl.styles.Font(italic=True)

# List of important item names
items = ["Progressive_Sword", "Progressive_Bow", "Progressive_Clawshot", "Progressive_Fishing_Rod", "Progressive_Sky_Book", "Progressive_Dominion_Rod", "Boomerang", "Lantern", "Spinner", "Ball_and_Chain", "Filled_Bomb_Bag" ,"Zora_Armor", "Gate_Key", "Shadow_Crytal" ]

# Dictionary to store woth_items of items
woth_items = {item: [] for item in items}

# Dictionary to store zones associated with each check
zones = {
"Arbiters Grounds Big Key Chest": "ArbitersGrounds",
"Arbiters Grounds Death Sword Chest": "ArbitersGrounds",
"Arbiters Grounds East Lower Turnable Redead Chest": "ArbitersGrounds",
"Arbiters Grounds East Turning Room Poe": "ArbitersGrounds",
"Arbiters Grounds East Upper Turnable Chest": "ArbitersGrounds",
"Arbiters Grounds East Upper Turnable Redead Chest": "ArbitersGrounds",
"Arbiters Grounds Entrance Chest": "ArbitersGrounds",
"Arbiters Grounds Ghoul Rat Room Chest": "ArbitersGrounds",
"Arbiters Grounds Hidden Wall Poe": "ArbitersGrounds",
"Arbiters Grounds North Turning Room Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room First Small Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Lower Central Small Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Lower North Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Second Small Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Stalfos Alcove Chest": "ArbitersGrounds",
"Arbiters Grounds Stallord Heart Container": "ArbitersGrounds",
"Arbiters Grounds Torch Room East Chest": "ArbitersGrounds",
"Arbiters Grounds Torch Room Poe": "ArbitersGrounds",
"Arbiters Grounds Torch Room West Chest": "ArbitersGrounds",
"Arbiters Grounds West Chandelier Chest": "ArbitersGrounds",
"Arbiters Grounds West Poe": "ArbitersGrounds",
"Arbiters Grounds West Small Chest Behind Block": "ArbitersGrounds",
"Arbiters Grounds West Stalfos Northeast Chest": "ArbitersGrounds",
"Arbiters Grounds West Stalfos West Chest": "ArbitersGrounds",
"Ashei Sketch": "SnowPeak Mountain",
"Barnes Bomb Bag": "Kakariko Village",
"Bridge of Eldin Female Phasmid": "North-East Hyrule Fields",
"Bridge of Eldin Male Phasmid": "Eldin Fields (East Hyrule Fields)",
"Bridge of Eldin Owl Statue Chest": "Eldin Fields (East Hyrule Fields)",
"Bridge of Eldin Owl Statue Sky Character": "Eldin Fields (East Hyrule Fields)",
"Bulblin Camp First Chest Under Tower At Entrance": "Bulbin Camp",
"Bulblin Camp Poe": "Bulbin Camp",
"Bulblin Camp Roasted Boar": "Bulbin Camp",
"Bulblin Camp Small Chest in Back of Camp": "Bulbin Camp",
"Bulblin Guard Key": "Bulbin Camp",
"Castle Town Malo Mart Magic Armor": "Hyrule Castle Town",
"Charlo Donation Blessing": "Hyrule Castle Town",
"City in The Sky Aeralfos Chest": "CityInTheSky",
"City in The Sky Argorok Heart Container": "CityInTheSky",
"City in The Sky Baba Tower Alcove Chest": "CityInTheSky",
"City in The Sky Baba Tower Narrow Ledge Chest": "CityInTheSky",
"City in The Sky Baba Tower Top Small Chest": "CityInTheSky",
"City in The Sky Big Key Chest": "CityInTheSky",
"City in The Sky Central Outside Ledge Chest": "CityInTheSky",
"City in The Sky Central Outside Poe Island Chest": "CityInTheSky",
"City in The Sky Chest Behind North Fan": "CityInTheSky",
"City in The Sky Chest Below Big Key Chest": "CityInTheSky",
"City in The Sky Dungeon Reward": "CityInTheSky",
"City in The Sky East First Wing Chest After Fans": "CityInTheSky",
"City in The Sky East Tile Worm Small Chest": "CityInTheSky",
"City in The Sky East Wing After Dinalfos Alcove Chest": "CityInTheSky",
"City in The Sky East Wing After Dinalfos Ledge Chest": "CityInTheSky",
"City in The Sky East Wing Lower Level Chest": "CityInTheSky",
"City in The Sky Garden Island Poe": "CityInTheSky",
"City in The Sky Poe Above Central Fan": "CityInTheSky",
"City in The Sky Underwater East Chest": "CityInTheSky",
"City in The Sky Underwater West Chest": "CityInTheSky",
"City in The Sky West Garden Corner Chest": "CityInTheSky",
"City in The Sky West Garden Ledge Chest": "CityInTheSky",
"City in The Sky West Garden Lone Island Chest": "CityInTheSky",
"City in The Sky West Garden Lower Chest": "CityInTheSky",
"City in The Sky West Wing Baba Balcony Chest": "CityInTheSky",
"City in The Sky West Wing First Chest": "CityInTheSky",
"City in The Sky West Wing Narrow Ledge Chest": "CityInTheSky",
"City in The Sky West Wing Tile Worm Chest": "CityInTheSky",
"Coro Bottle": "Faron Woods",
"Death Mountain Alcove Chest": "Death Mountain",
"Death Mountain Trail Poe": "Death Mountain",
"Doctors Office Balcony Chest": "Hyrule Castle Town",
"East Castle Town Bridge Poe": "Hyrule Castle Town",
"Eldin Field Bomb Rock Chest": "Eldin Fields (East Hyrule Fields)",
"Eldin Field Bomskit Grotto Lantern Chest": "Eldin Fields (East Hyrule Fields)",
"Eldin Field Bomskit Grotto Left Chest": "Eldin Fields (East Hyrule Fields)",
"Eldin Field Female Grasshopper": "Eldin Fields (East Hyrule Fields)",
"Eldin Field Male Grasshopper": "Eldin Fields (East Hyrule Fields)",
"Eldin Field Stalfos Grotto Left Small Chest": "North-East Hyrule Fields",
"Eldin Field Stalfos Grotto Right Small Chest": "North-East Hyrule Fields",
"Eldin Field Stalfos Grotto Stalfos Chest": "North-East Hyrule Fields",
"Eldin Field Water Bomb Fish Grotto Chest": "Eldin Fields (East Hyrule Fields)",
"Eldin Lantern Cave First Chest": "Eldin Gorge",
"Eldin Lantern Cave Lantern Chest": "Eldin Gorge",
"Eldin Lantern Cave Poe": "Eldin Gorge",
"Eldin Lantern Cave Second Chest": "Eldin Gorge",
"Eldin Spring Underwater Chest": "Kakariko Village",
"Eldin Stockcave Lantern Chest": "North-East Hyrule Fields",
"Eldin Stockcave Lowest Chest": "North-East Hyrule Fields",
"Eldin Stockcave Upper Chest": "North-East Hyrule Fields",
"Faron Field Bridge Chest": "South Hyrule Fields",
"Faron Field Corner Grotto Left Chest": "South Hyrule Fields",
"Faron Field Corner Grotto Rear Chest": "South Hyrule Fields",
"Faron Field Corner Grotto Right Chest": "South Hyrule Fields",
"Faron Field Female Beetle": "South Hyrule Fields",
"Faron Field Male Beetle": "South Hyrule Fields",
"Faron Field Poe": "South Hyrule Fields",
"Faron Field Tree Heart Piece": "South Hyrule Fields",
"Faron Mist Cave Lantern Chest": "Faron Woods",
"Faron Mist Cave Open Chest": "Faron Woods",
"Faron Mist North Chest": "Faron Woods",
"Faron Mist Poe": "Faron Woods",
"Faron Mist South Chest": "Faron Woods",
"Faron Mist Stump Chest": "Faron Woods",
"Faron Woods Golden Wolf": "Faron Woods",
"Faron Woods Owl Statue Chest": "Faron Woods",
"Faron Woods Owl Statue Sky Character": "Faron Woods",
"Fishing Hole Bottle": "Upper Zora's River",
"Fishing Hole Heart Piece": "Upper Zora's River",
"Flight By Fowl Fifth Platform Chest": "Lake Hylia",
"Flight By Fowl Fourth Platform Chest": "Lake Hylia",
"Flight By Fowl Ledge Poe": "Lake Hylia",
"Flight By Fowl Second Platform Chest": "Lake Hylia",
"Flight By Fowl Third Platform Chest": "Lake Hylia",
"Flight By Fowl Top Platform Reward": "Lake Hylia",
"Forest Temple Big Baba Key": "ForestTemple",
"Forest Temple Big Key Chest": "ForestTemple",
"Forest Temple Central Chest Behind Stairs": "ForestTemple",
"Forest Temple Central Chest Hanging From Web": "ForestTemple",
"Forest Temple Central North Chest": "ForestTemple",
"Forest Temple Diababa Heart Container": "ForestTemple",
"Forest Temple Dungeon Reward": "ForestTemple",
"Forest Temple East Tile Worm Chest": "ForestTemple",
"Forest Temple East Water Cave Chest": "ForestTemple",
"Forest Temple Entrance Vines Chest": "ForestTemple",
"Forest Temple Gale Boomerang": "ForestTemple",
"Forest Temple North Deku Like Chest": "ForestTemple",
"Forest Temple Second Monkey Under Bridge Chest": "ForestTemple",
"Forest Temple Totem Pole Chest": "ForestTemple",
"Forest Temple West Deku Like Chest": "ForestTemple",
"Forest Temple West Tile Worm Chest Behind Stairs": "ForestTemple",
"Forest Temple West Tile Worm Room Vines Chest": "ForestTemple",
"Forest Temple Windless Bridge Chest": "ForestTemple",
"Gerudo Desert Campfire East Chest": "Gerudo Desert",
"Gerudo Desert Campfire North Chest": "Gerudo Desert",
"Gerudo Desert Campfire West Chest": "Gerudo Desert",
"Gerudo Desert East Canyon Chest": "Gerudo Desert",
"Gerudo Desert East Poe": "Gerudo Desert",
"Gerudo Desert Female Dayfly": "Gerudo Desert",
"Gerudo Desert Golden Wolf": "Gerudo Desert",
"Gerudo Desert Lone Small Chest": "Gerudo Desert",
"Gerudo Desert Male Dayfly": "Gerudo Desert",
"Gerudo Desert North Peahat Poe": "Gerudo Desert",
"Gerudo Desert North Small Chest Before Bulblin Camp": "Gerudo Desert",
"Gerudo Desert Northeast Chest Behind Gates": "Gerudo Desert",
"Gerudo Desert Northwest Chest Behind Gates": "Gerudo Desert",
"Gerudo Desert Owl Statue Chest": "Gerudo Desert",
"Gerudo Desert Owl Statue Sky Character": "Gerudo Desert",
"Gerudo Desert Peahat Ledge Chest": "Gerudo Desert",
"Gerudo Desert Poe Above Cave of Ordeals": "Gerudo Desert",
"Gerudo Desert Rock Grotto First Poe": "Gerudo Desert",
"Gerudo Desert Rock Grotto Lantern Chest": "Gerudo Desert",
"Gerudo Desert Rock Grotto Second Poe": "Gerudo Desert",
"Gerudo Desert Skulltula Grotto Chest": "Gerudo Desert",
"Gerudo Desert South Chest Behind Wooden Gates": "Gerudo Desert",
"Gerudo Desert West Canyon Chest": "Gerudo Desert",
"Gift From Ralis": "Kakariko Graveyard",
"Goron Mines After Crystal Switch Room Magnet Wall Chest": "GoronMines",
"Goron Mines Beamos Room Chest": "GoronMines",
"Goron Mines Chest Before Dangoro": "GoronMines",
"Goron Mines Crystal Switch Room Small Chest": "GoronMines",
"Goron Mines Crystal Switch Room Underwater Chest": "GoronMines",
"Goron Mines Dangoro Chest": "GoronMines",
"Goron Mines Dungeon Reward": "GoronMines",
"Goron Mines Entrance Chest": "GoronMines",
"Goron Mines Fyrus Heart Container": "GoronMines",
"Goron Mines Gor Amato Chest": "GoronMines",
"Goron Mines Gor Amato Key Shard": "GoronMines",
"Goron Mines Gor Amato Small Chest": "GoronMines",
"Goron Mines Gor Ebizo Chest": "GoronMines",
"Goron Mines Gor Ebizo Key Shard": "GoronMines",
"Goron Mines Gor Liggs Chest": "GoronMines",
"Goron Mines Gor Liggs Key Shard": "GoronMines",
"Goron Mines Magnet Maze Chest": "GoronMines",
"Goron Mines Main Magnet Room Bottom Chest": "GoronMines",
"Goron Mines Main Magnet Room Top Chest": "GoronMines",
"Goron Mines Outside Beamos Chest": "GoronMines",
"Goron Mines Outside Clawshot Chest": "GoronMines",
"Goron Mines Outside Underwater Chest": "GoronMines",
"Goron Springwater Rush": "Eldin Fields (East Hyrule Fields)",
"Herding Goats Reward": "Ordon Village",
"Hyrule Castle Big Key Chest": "Hyrule Castle",
"Hyrule Castle East Wing Balcony Chest": "Hyrule Castle",
"Hyrule Castle East Wing Boomerang Puzzle Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Grave Switch Room Back Left Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Grave Switch Room Front Left Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Grave Switch Room Right Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Owl Statue Chest": "Hyrule Castle",
"Hyrule Castle King Bulblin Key": "Hyrule Castle",
"Hyrule Castle Lantern Staircase Chest": "HyruleCastle",
"Hyrule Castle Main Hall Northeast Chest": "HyruleCastle",
"Hyrule Castle Main Hall Northwest Chest": "HyruleCastle",
"Hyrule Castle Main Hall Southwest Chest": "HyruleCastle",
"Hyrule Castle Southeast Balcony Tower Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Eighth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fifth Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fifth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room First Chest": "HyruleCastle",
"Hyrule Castle Treasure Room First Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fourth Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fourth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Second Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Second Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Seventh Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Sixth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Third Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Third Small Chest": "HyruleCastle",
"Hyrule Castle West Courtyard Central Small Chest": "HyruleCastle",
"Hyrule Castle West Courtyard North Small Chest": "HyruleCastle",
"Hyrule Field Ampitheater Owl Statue Chest": "Outside West Castle Town (West Fields)",
"Hyrule Field Ampitheater Owl Statue Sky Character": "Outside West Castle Town (West Fields)",
"Hyrule Field Ampitheater Poe": "Outside West Castle Town (West Fields)",
"Isle of Riches Poe": "Lake Hylia",
"Iza Helping Hand": "Lake Hylia",
"Iza Raging Rapids Minigame": "Lake Hylia",
"Jovani 20 Poe Soul Reward": "Hyrule Castle Town",
"Jovani 60 Poe Soul Reward": "Hyrule Castle Town",
"Jovani House Poe": "Hyrule Castle Town",
"Kakariko Gorge Double Clawshot Chest": "Eldin Gorge",
"Kakariko Gorge Female Pill Bug": "Eldin Gorge",
"Kakariko Gorge Male Pill Bug": "Eldin Gorge",
"Kakariko Gorge Owl Statue Chest": "Eldin Gorge",
"Kakariko Gorge Owl Statue Sky Character": "Eldin Gorge",
"Kakariko Gorge Poe": "Eldin Gorge",
"Kakariko Gorge Spire Heart Piece": "Eldin Gorge",
"Kakariko Graveyard Golden Wolf": "Kakariko Graveyard",
"Kakariko Graveyard Grave Poe": "Kakariko Graveyard",
"Kakariko Graveyard Lantern Chest": "Kakariko Graveyard",
"Kakariko Graveyard Male Ant": "Kakariko Graveyard",
"Kakariko Graveyard Open Poe": "Kakariko Graveyard",
"Kakariko Inn Chest": "Kakariko Village",
"Kakariko Village Bomb Rock Spire Heart Piece": "Kakariko Village",
"Kakariko Village Bomb Shop Poe": "Kakariko Village",
"Kakariko Village Female Ant": "Kakariko Village",
"Kakariko Village Malo Mart Hawkeye": "Kakariko Village",
"Kakariko Village Malo Mart Hylian Shield": "Kakariko Village",
"Kakariko Village Watchtower Poe": "Kakariko Village",
"Kakariko Watchtower Alcove Chest": "Kakariko Village",
"Kakariko Watchtower Chest": "Kakariko Village",
"Lake Hylia Alcove Poe": "Lake Hylia",
"Lake Hylia Bridge Bubble Grotto Chest": "Great Bridge of Hylia",
"Lake Hylia Bridge Cliff Chest": "Great Bridge of Hylia",
"Lake Hylia Bridge Cliff Poe": "Great Bridge of Hylia",
"Lake Hylia Bridge Female Mantis": "Great Bridge of Hylia",
"Lake Hylia Bridge Male Mantis": "Great Bridge of Hylia",
"Lake Hylia Bridge Owl Statue Chest": "Great Bridge of Hylia",
"Lake Hylia Bridge Owl Statue Sky Character": "Great Bridge of Hylia",
"Lake Hylia Bridge Vines Chest": "Great Bridge of Hylia",
"Lake Hylia Dock Poe": "Lake Hylia",
"Lake Hylia Shell Blade Grotto Chest": "Lake Hylia",
"Lake Hylia Tower Poe": "Lake Hylia",
"Lake Hylia Underwater Chest": "Lake Hylia",
"Lake Hylia Water Toadpoli Grotto Chest": "Lake Hylia",
"Lake Lantern Cave Eighth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Eleventh Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave End Lantern Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Fifth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Final Poe": "Lake Hylia Lantern Cave",
"Lake Lantern Cave First Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave First Poe": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Fourteenth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Fourth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Ninth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Second Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Second Poe": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Seventh Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Sixth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Tenth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Third Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Thirteenth Chest": "Lake Hylia Lantern Cave",
"Lake Lantern Cave Twelfth Chest": "Lake Hylia Lantern Cave",
"Lakebed Temple Before Deku Toad Alcove Chest": "LakebedTemple",
"Lakebed Temple Before Deku Toad Underwater Left Chest": "LakebedTemple",
"Lakebed Temple Before Deku Toad Underwater Right Chest": "LakebedTemple",
"Lakebed Temple Big Key Chest": "LakebedTemple",
"Lakebed Temple Central Room Chest": "LakebedTemple",
"Lakebed Temple Central Room Small Chest": "LakebedTemple",
"Lakebed Temple Central Room Spire Chest": "LakebedTemple",
"Lakebed Temple Chandelier Chest": "LakebedTemple",
"Lakebed Temple Deku Toad Chest": "LakebedTemple",
"Lakebed Temple Dungeon Reward": "LakebedTemple",
"Lakebed Temple East Lower Waterwheel Bridge Chest": "LakebedTemple",
"Lakebed Temple East Lower Waterwheel Stalactite Chest": "LakebedTemple",
"Lakebed Temple East Second Floor Southeast Chest": "LakebedTemple",
"Lakebed Temple East Second Floor Southwest Chest": "LakebedTemple",
"Lakebed Temple East Water Supply Clawshot Chest": "LakebedTemple",
"Lakebed Temple East Water Supply Small Chest": "LakebedTemple",
"Lakebed Temple Lobby Left Chest": "LakebedTemple",
"Lakebed Temple Lobby Rear Chest": "LakebedTemple",
"Lakebed Temple Morpheel Heart Container": "LakebedTemple",
"Lakebed Temple Stalactite Room Chest": "LakebedTemple",
"Lakebed Temple Underwater Maze Small Chest": "LakebedTemple",
"Lakebed Temple West Lower Small Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Central Small Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Northeast Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Southeast Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Southwest Underwater Chest": "LakebedTemple",
"Lakebed Temple West Water Supply Chest": "LakebedTemple",
"Lakebed Temple West Water Supply Small Chest": "LakebedTemple",
"Lanayru Field Behind Gate Underwater Chest": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Bridge Poe": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Female Stag Beetle": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Male Stag Beetle": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Poe Grotto Left Poe": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Poe Grotto Right Poe": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Skulltula Grotto Chest": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Field Spinner Track Chest": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Ice Block Puzzle Cave Chest": "Lanayru Fields (North Hyrule Fields)",
"Lanayru Spring Back Room Lantern Chest": "Lanayru Spring",
"Lanayru Spring Back Room Left Chest": "Lanayru Spring",
"Lanayru Spring Back Room Right Chest": "Lanayru Spring",
"Lanayru Spring East Double Clawshot Chest": "Lanayru Spring",
"Lanayru Spring Underwater Left Chest": "Lanayru Spring",
"Lanayru Spring Underwater Right Chest": "Lanayru Spring",
"Lanayru Spring West Double Clawshot Chest": "Lanayru Spring",
"Links Basement Chest": "Ordon Village",
"Lost Woods Boulder Poe": "Sacred Grove & Lost Woods",
"Lost Woods Lantern Chest": "Sacred Grove & Lost Woods",
"Lost Woods Waterfall Poe": "Sacred Grove & Lost Woods",
"North Castle Town Golden Wolf": "Hyrule Castle Town",
"North Faron Woods Deku Baba Chest": "Faron Woods",
"Ordon Ranch Grotto Lantern Chest": "Ordon Village",
"Ordon Spring Golden Wolf": "Ordon Village",
"Outside Arbiters Grounds Lantern Chest": "Bulbin Camp",
"Outside Arbiters Grounds Poe": "Bulbin Camp",
"Outside Bulblin Camp Poe": "Bulbin Camp",
"Outside Lanayru Spring Left Statue Chest": "Lake Hylia",
"Outside Lanayru Spring Right Statue Chest": "Lake Hylia",
"Outside South Castle Town Double Clawshot Chasm Chest": "South Castle Town",
"Outside South Castle Town Female Ladybug": "South Castle Town",
"Outside South Castle Town Fountain Chest": "South Castle Town",
"Outside South Castle Town Golden Wolf": "South Castle Town",
"Outside South Castle Town Male Ladybug": "South Castle Town",
"Outside South Castle Town Poe": "South Castle Town",
"Outside South Castle Town Tektite Grotto Chest": "South Castle Town",
"Outside South Castle Town Tightrope Chest": "South Castle Town",
"Palace of Twilight Big Key Chest": "PalaceOfTwilight",
"Palace of Twilight Central First Room Chest": "PalaceOfTwilight",
"Palace of Twilight Central Outdoor Chest": "PalaceOfTwilight",
"Palace of Twilight Central Tower Chest": "PalaceOfTwilight",
"Palace of Twilight Collect Both Sols": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room East Alcove": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room North Small Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room West Alcove": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room Zant Head Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Northeast Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Northwest Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Southeast Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Southwest Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Chest Behind Wall of Darkness": "PalaceOfTwilight",
"Palace of Twilight West Wing First Room Central Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Second Room Central Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Second Room Lower South Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Second Room Southeast Chest": "PalaceOfTwilight",
"Palace of Twilight Zant Heart Container": "PalaceOfTwilight",
"Plumm Fruit Balloon Minigame": "Upper Zora’s River",
"STAR Prize 1": "Hyrule Castle Town",
"STAR Prize 2": "Hyrule Castle Town",
"Sacred Grove Baba Serpent Grotto Chest": "Sacred Grove & Lost Woods",
"Sacred Grove Female Snail": "Sacred Grove & Lost Woods",
"Sacred Grove Male Snail": "Sacred Grove & Lost Woods",
"Sacred Grove Master Sword Poe": "Sacred Grove & Lost Woods",
"Sacred Grove Past Owl Statue Chest": "Sacred Grove & Lost Woods",
"Sacred Grove Spinner Chest": "Sacred Grove & Lost Woods",
"Sacred Grove Temple of Time Owl Statue Poe": "Sacred Grove & Lost Woods",
"Sera Shop Slingshot": "Ordon Village",
"Snowboard Racing Prize": "SnowPeak Mountain",
"Snowpeak Above Freezard Grotto Poe": "SnowPeak Mountain",
"Snowpeak Blizzard Poe": "SnowPeak Mountain",
"Snowpeak Cave Ice Lantern Chest": "SnowPeak Mountain",
"Snowpeak Cave Ice Poe": "SnowPeak Mountain",
"Snowpeak Freezard Grotto Chest": "SnowPeak Mountain",
"Snowpeak Icy Summit Poe": "SnowPeak Mountain",
"Snowpeak Poe Among Trees": "SnowPeak Mountain",
"Snowpeak Ruins Ball and Chain": "SnowpeakRuins",
"Snowpeak Ruins Blizzeta Heart Container": "SnowpeakRuins",
"Snowpeak Ruins Broken Floor Chest": "SnowpeakRuins",
"Snowpeak Ruins Chapel Chest": "SnowpeakRuins",
"Snowpeak Ruins Chest After Darkhammer": "SnowpeakRuins",
"Snowpeak Ruins Courtyard Central Chest": "SnowpeakRuins",
"Snowpeak Ruins Dungeon Reward": "SnowpeakRuins",
"Snowpeak Ruins East Courtyard Buried Chest": "SnowpeakRuins",
"Snowpeak Ruins East Courtyard Chest": "SnowpeakRuins",
"Snowpeak Ruins Ice Room Poe": "SnowpeakRuins",
"Snowpeak Ruins Lobby Armor Poe": "SnowpeakRuins",
"Snowpeak Ruins Lobby Chandelier Chest": "SnowpeakRuins",
"Snowpeak Ruins Lobby East Armor Chest": "SnowpeakRuins",
"Snowpeak Ruins Lobby Poe": "SnowpeakRuins",
"Snowpeak Ruins Lobby West Armor Chest": "SnowpeakRuins",
"Snowpeak Ruins Mansion Map": "SnowpeakRuins",
"Snowpeak Ruins Northeast Chandelier Chest": "SnowpeakRuins",
"Snowpeak Ruins Ordon Pumpkin Chest": "SnowpeakRuins",
"Snowpeak Ruins West Cannon Room Central Chest": "SnowpeakRuins",
"Snowpeak Ruins West Cannon Room Corner Chest": "SnowpeakRuins",
"Snowpeak Ruins West Courtyard Buried Chest": "SnowpeakRuins",
"Snowpeak Ruins Wooden Beam Central Chest": "SnowpeakRuins",
"Snowpeak Ruins Wooden Beam Chandelier Chest": "SnowpeakRuins",
"Snowpeak Ruins Wooden Beam Northwest Chest": "SnowpeakRuins",
"South Faron Cave Chest": "Faron Woods",
"Talo Sharpshooting": "Kakariko Village",
"Temple of Time Armogohma Heart Container": "TempleOfTime",
"Temple of Time Armos Antechamber East Chest": "TempleOfTime",
"Temple of Time Armos Antechamber North Chest": "TempleOfTime",
"Temple of Time Armos Antechamber Statue Chest": "TempleOfTime",
"Temple of Time Big Key Chest": "TempleOfTime",
"Temple of Time Chest Before Darknut": "TempleOfTime",
"Temple of Time Darknut Chest": "TempleOfTime",
"Temple of Time Dungeon Reward": "TempleOfTime",
"Temple of Time First Staircase Armos Chest": "TempleOfTime",
"Temple of Time First Staircase Gohma Gate Chest": "TempleOfTime",
"Temple of Time First Staircase Window Chest": "TempleOfTime",
"Temple of Time Floor Switch Puzzle Room Upper Chest": "TempleOfTime",
"Temple of Time Gilloutine Chest": "TempleOfTime",
"Temple of Time Lobby Lantern Chest": "TempleOfTime",
"Temple of Time Moving Wall Beamos Room Chest": "TempleOfTime",
"Temple of Time Moving Wall Dinalfos Room Chest": "TempleOfTime",
"Temple of Time Poe Above Scales": "TempleOfTime",
"Temple of Time Poe Behind Gate": "TempleOfTime",
"Temple of Time Scales Gohma Chest": "TempleOfTime",
"Temple of Time Scales Upper Chest": "TempleOfTime",
"Upper Zoras River Female Dragonfly": "Upper Zora’s River",
"Upper Zoras River Poe": "Upper Zora’s River",
"West Hyrule Field Female Butterfly": "Outside West Castle Town (West Fields)",
"West Hyrule Field Golden Wolf": "Outside West Castle Town (West Fields)",
"West Hyrule Field Helmasaur Grotto Chest": "Outside West Castle Town (West Fields)",
"West Hyrule Field Male Butterfly": "Outside West Castle Town (West Fields)",
"Wooden Sword Chest": "Ordon Village",
"Wrestling With Bo": "Ordon Village",
"Zoras Domain Chest Behind Waterfall": "Zoras Domain",
"Zoras Domain Chest By Mother and Child Isles": "Zoras Domain",
"Zoras Domain Extinguish All Torches Chest": "Zoras Domain",
"Zoras Domain Light All Torches Chest": "Zoras Domain",
"Zoras Domain Male Dragonfly": "Zoras Domain",
"Zoras Domain Mother and Child Isle Poe": "Zoras Domain",
"Zoras Domain Underwater Goron": "Zoras Domain",
"Zoras Domain Waterfall Poe": "Zoras Domain",
}

# List to store WotH phrases
important_zones = []

# List to store WotH zones
woth_zone_list = []

# List to store Barren zones
unimportant_zones = []

# Dungeon list
dungeons = ["ForestTemple", "GoronMines", "LakebedTemple", "ArbitersGrounds", "SnowpeakRuins", "TempleOfTime", "CityInTheSky", "PalaceOfTwilight"]

# Get the list of required dungeons from the json file
required_dungeons = data["requiredDungeons"]

# Iterate through the spheres in the json file
for sphere_name, sphere in data["spheres"].items():
    # Iterate through the checks in each sphere
    for check_name, item in sphere.items():
        # If the item is in the item list, add the associated checks to the good group
        if item in items:
            woth_items[item].extend([check_name])

# Define WotH zones
for item, check_list in woth_items.items():
    for check in check_list:
        zone = "unknown"
        if check in zones:
            zone = zones[check]
            if zone not in important_zones:
                important_zones.append(zone)
                woth_zone_list.append(zone)

# Select 3 random WotH
selected_zones = random.sample(important_zones, 3)
with open('Hint_list.txt', "w") as file:
    for zone in selected_zones:
        file.write(zone + ' is on the way of the hero\n')

# Add WotH to the xlsx file
for i in range(3, 6):
    zone = selected_zones[i-3]
    worksheet.cell(row=i, column=2, value=zone)
    worksheet.cell(row=i, column=2, value=zone + " is on the way of the Hero")
    
# Prompt WotH
for zone in selected_zones:
    print("{} is on the way of the Hero".format(zone))

# Stock all zone names
unique_zone_names = set()

# List all zone names
for check, zone in zones.items():
    unique_zone_names.add(zone)

# Barren zones definition process
barren = unique_zone_names
barren = barren - set(woth_zone_list)

# Exclude from the dungeons list the ones that are in the required dungeons list
required_dungeons = data["requiredDungeons"]
for dungeon in required_dungeons:
    if dungeon in dungeons:
        dungeons.remove(dungeon)

# Remove from the barren zones the not required dungeons
barren = barren - set(dungeons)

# Select 3 random barren
selected_zones = random.sample(list(barren), 2)
with open('Hint_list.txt', "a") as file:
    for zone in selected_zones:
        file.write(zone + ' is barren\n')        
# Add barren to the xlsx file
for i in range(8, 10):
    zone = selected_zones[i-8]
    worksheet.cell(row=i, column=2, value=zone)
    worksheet.cell(row=i, column=2, value=zone + " is barren")
    
# Prompt barren
for zone in selected_zones:
    print(f"{zone} is barren")

# Always_checks list
required_checks = ["Jovani 20 Poe Soul Reward", "Goron Springwater Rush", "Lanayru Ice Block Puzzle Cave Chest", "Iza Helping Hand"]

# Dictionary to store checks and item linked to those checks
always_items = {}

# Analyse Always checks
for check, item in data["itemPlacements"].items():
    if check in required_checks:
        always_items[check] = item

# Add Always to the xlsx file
for i, (check, item) in enumerate(always_items.items(), start=12):
    worksheet.cell(row=i, column=2, value=f'{check} : {item}')

# Prompt Always
for check, item in always_items.items():
    # Afficher la phrase "nom du check" : "item"
    print(f'{check} : {item}')
    with open('Hint_list.txt', "a") as file:
        file.write(f'{check} : {item}\n')

# Dungeons checks list with their zones
all_dungeon_checks = {
"Arbiters Grounds Big Key Chest": "ArbitersGrounds",
"Arbiters Grounds Death Sword Chest": "ArbitersGrounds",
"Arbiters Grounds East Lower Turnable Redead Chest": "ArbitersGrounds",
"Arbiters Grounds East Turning Room Poe": "ArbitersGrounds",
"Arbiters Grounds East Upper Turnable Chest": "ArbitersGrounds",
"Arbiters Grounds East Upper Turnable Redead Chest": "ArbitersGrounds",
"Arbiters Grounds Entrance Chest": "ArbitersGrounds",
"Arbiters Grounds Ghoul Rat Room Chest": "ArbitersGrounds",
"Arbiters Grounds Hidden Wall Poe": "ArbitersGrounds",
"Arbiters Grounds North Turning Room Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room First Small Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Lower Central Small Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Lower North Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Second Small Chest": "ArbitersGrounds",
"Arbiters Grounds Spinner Room Stalfos Alcove Chest": "ArbitersGrounds",
"Arbiters Grounds Stallord Heart Container": "ArbitersGrounds",
"Arbiters Grounds Torch Room East Chest": "ArbitersGrounds",
"Arbiters Grounds Torch Room Poe": "ArbitersGrounds",
"Arbiters Grounds Torch Room West Chest": "ArbitersGrounds",
"Arbiters Grounds West Chandelier Chest": "ArbitersGrounds",
"Arbiters Grounds West Poe": "ArbitersGrounds",
"Arbiters Grounds West Small Chest Behind Block": "ArbitersGrounds",
"Arbiters Grounds West Stalfos Northeast Chest": "ArbitersGrounds",
"Arbiters Grounds West Stalfos West Chest": "ArbitersGrounds",
"City in The Sky Aeralfos Chest": "CityInTheSky",
"City in The Sky Argorok Heart Container": "CityInTheSky",
"City in The Sky Baba Tower Alcove Chest": "CityInTheSky",
"City in The Sky Baba Tower Narrow Ledge Chest": "CityInTheSky",
"City in The Sky Baba Tower Top Small Chest": "CityInTheSky",
"City in The Sky Big Key Chest": "CityInTheSky",
"City in The Sky Central Outside Ledge Chest": "CityInTheSky",
"City in The Sky Central Outside Poe Island Chest": "CityInTheSky",
"City in The Sky Chest Behind North Fan": "CityInTheSky",
"City in The Sky Chest Below Big Key Chest": "CityInTheSky",
"City in The Sky Dungeon Reward": "CityInTheSky",
"City in The Sky East First Wing Chest After Fans": "CityInTheSky",
"City in The Sky East Tile Worm Small Chest": "CityInTheSky",
"City in The Sky East Wing After Dinalfos Alcove Chest": "CityInTheSky",
"City in The Sky East Wing After Dinalfos Ledge Chest": "CityInTheSky",
"City in The Sky East Wing Lower Level Chest": "CityInTheSky",
"City in The Sky Garden Island Poe": "CityInTheSky",
"City in The Sky Poe Above Central Fan": "CityInTheSky",
"City in The Sky Underwater East Chest": "CityInTheSky",
"City in The Sky Underwater West Chest": "CityInTheSky",
"City in The Sky West Garden Corner Chest": "CityInTheSky",
"City in The Sky West Garden Ledge Chest": "CityInTheSky",
"City in The Sky West Garden Lone Island Chest": "CityInTheSky",
"City in The Sky West Garden Lower Chest": "CityInTheSky",
"City in The Sky West Wing Baba Balcony Chest": "CityInTheSky",
"City in The Sky West Wing First Chest": "CityInTheSky",
"City in The Sky West Wing Narrow Ledge Chest": "CityInTheSky",
"City in The Sky West Wing Tile Worm Chest": "CityInTheSky",
"Forest Temple Big Baba Key": "ForestTemple",
"Forest Temple Big Key Chest": "ForestTemple",
"Forest Temple Central Chest Behind Stairs": "ForestTemple",
"Forest Temple Central Chest Hanging From Web": "ForestTemple",
"Forest Temple Central North Chest": "ForestTemple",
"Forest Temple Diababa Heart Container": "ForestTemple",
"Forest Temple Dungeon Reward": "ForestTemple",
"Forest Temple East Tile Worm Chest": "ForestTemple",
"Forest Temple East Water Cave Chest": "ForestTemple",
"Forest Temple Entrance Vines Chest": "ForestTemple",
"Forest Temple Gale Boomerang": "ForestTemple",
"Forest Temple North Deku Like Chest": "ForestTemple",
"Forest Temple Second Monkey Under Bridge Chest": "ForestTemple",
"Forest Temple Totem Pole Chest": "ForestTemple",
"Forest Temple West Deku Like Chest": "ForestTemple",
"Forest Temple West Tile Worm Chest Behind Stairs": "ForestTemple",
"Forest Temple West Tile Worm Room Vines Chest": "ForestTemple",
"Forest Temple Windless Bridge Chest": "ForestTemple",
"Goron Mines After Crystal Switch Room Magnet Wall Chest": "GoronMines",
"Goron Mines Beamos Room Chest": "GoronMines",
"Goron Mines Chest Before Dangoro": "GoronMines",
"Goron Mines Crystal Switch Room Small Chest": "GoronMines",
"Goron Mines Crystal Switch Room Underwater Chest": "GoronMines",
"Goron Mines Dangoro Chest": "GoronMines",
"Goron Mines Dungeon Reward": "GoronMines",
"Goron Mines Entrance Chest": "GoronMines",
"Goron Mines Fyrus Heart Container": "GoronMines",
"Goron Mines Gor Amato Chest": "GoronMines",
"Goron Mines Gor Amato Key Shard": "GoronMines",
"Goron Mines Gor Amato Small Chest": "GoronMines",
"Goron Mines Gor Ebizo Chest": "GoronMines",
"Goron Mines Gor Ebizo Key Shard": "GoronMines",
"Goron Mines Gor Liggs Chest": "GoronMines",
"Goron Mines Gor Liggs Key Shard": "GoronMines",
"Goron Mines Magnet Maze Chest": "GoronMines",
"Goron Mines Main Magnet Room Bottom Chest": "GoronMines",
"Goron Mines Main Magnet Room Top Chest": "GoronMines",
"Goron Mines Outside Beamos Chest": "GoronMines",
"Goron Mines Outside Clawshot Chest": "GoronMines",
"Goron Mines Outside Underwater Chest": "GoronMines",
"Hyrule Castle Big Key Chest": "Hyrule Castle",
"Hyrule Castle East Wing Balcony Chest": "Hyrule Castle",
"Hyrule Castle East Wing Boomerang Puzzle Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Grave Switch Room Back Left Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Grave Switch Room Front Left Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Grave Switch Room Right Chest": "Hyrule Castle",
"Hyrule Castle Graveyard Owl Statue Chest": "Hyrule Castle",
"Hyrule Castle King Bulblin Key": "Hyrule Castle",
"Hyrule Castle Lantern Staircase Chest": "HyruleCastle",
"Hyrule Castle Main Hall Northeast Chest": "HyruleCastle",
"Hyrule Castle Main Hall Northwest Chest": "HyruleCastle",
"Hyrule Castle Main Hall Southwest Chest": "HyruleCastle",
"Hyrule Castle Southeast Balcony Tower Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Eighth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fifth Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fifth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room First Chest": "HyruleCastle",
"Hyrule Castle Treasure Room First Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fourth Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Fourth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Second Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Second Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Seventh Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Sixth Small Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Third Chest": "HyruleCastle",
"Hyrule Castle Treasure Room Third Small Chest": "HyruleCastle",
"Hyrule Castle West Courtyard Central Small Chest": "HyruleCastle",
"Hyrule Castle West Courtyard North Small Chest": "HyruleCastle",
"Lakebed Temple Before Deku Toad Alcove Chest": "LakebedTemple",
"Lakebed Temple Before Deku Toad Underwater Left Chest": "LakebedTemple",
"Lakebed Temple Before Deku Toad Underwater Right Chest": "LakebedTemple",
"Lakebed Temple Big Key Chest": "LakebedTemple",
"Lakebed Temple Central Room Chest": "LakebedTemple",
"Lakebed Temple Central Room Small Chest": "LakebedTemple",
"Lakebed Temple Central Room Spire Chest": "LakebedTemple",
"Lakebed Temple Chandelier Chest": "LakebedTemple",
"Lakebed Temple Deku Toad Chest": "LakebedTemple",
"Lakebed Temple Dungeon Reward": "LakebedTemple",
"Lakebed Temple East Lower Waterwheel Bridge Chest": "LakebedTemple",
"Lakebed Temple East Lower Waterwheel Stalactite Chest": "LakebedTemple",
"Lakebed Temple East Second Floor Southeast Chest": "LakebedTemple",
"Lakebed Temple East Second Floor Southwest Chest": "LakebedTemple",
"Lakebed Temple East Water Supply Clawshot Chest": "LakebedTemple",
"Lakebed Temple East Water Supply Small Chest": "LakebedTemple",
"Lakebed Temple Lobby Left Chest": "LakebedTemple",
"Lakebed Temple Lobby Rear Chest": "LakebedTemple",
"Lakebed Temple Morpheel Heart Container": "LakebedTemple",
"Lakebed Temple Stalactite Room Chest": "LakebedTemple",
"Lakebed Temple Underwater Maze Small Chest": "LakebedTemple",
"Lakebed Temple West Lower Small Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Central Small Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Northeast Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Southeast Chest": "LakebedTemple",
"Lakebed Temple West Second Floor Southwest Underwater Chest": "LakebedTemple",
"Lakebed Temple West Water Supply Chest": "LakebedTemple",
"Lakebed Temple West Water Supply Small Chest": "LakebedTemple",
"Palace of Twilight Big Key Chest": "PalaceOfTwilight",
"Palace of Twilight Central First Room Chest": "PalaceOfTwilight",
"Palace of Twilight Central Outdoor Chest": "PalaceOfTwilight",
"Palace of Twilight Central Tower Chest": "PalaceOfTwilight",
"Palace of Twilight Collect Both Sols": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room East Alcove": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room North Small Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room West Alcove": "PalaceOfTwilight",
"Palace of Twilight East Wing First Room Zant Head Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Northeast Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Northwest Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Southeast Chest": "PalaceOfTwilight",
"Palace of Twilight East Wing Second Room Southwest Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Chest Behind Wall of Darkness": "PalaceOfTwilight",
"Palace of Twilight West Wing First Room Central Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Second Room Central Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Second Room Lower South Chest": "PalaceOfTwilight",
"Palace of Twilight West Wing Second Room Southeast Chest": "PalaceOfTwilight",
"Palace of Twilight Zant Heart Container": "PalaceOfTwilight",
"Snowpeak Above Freezard Grotto Poe": "SnowPeak Mountain",
"Snowpeak Blizzard Poe": "SnowPeak Mountain",
"Snowpeak Cave Ice Lantern Chest": "SnowPeak Mountain",
"Snowpeak Cave Ice Poe": "SnowPeak Mountain",
"Snowpeak Freezard Grotto Chest": "SnowPeak Mountain",
"Snowpeak Icy Summit Poe": "SnowPeak Mountain",
"Snowpeak Poe Among Trees": "SnowPeak Mountain",
"Snowpeak Ruins Ball and Chain": "SnowpeakRuins",
"Snowpeak Ruins Blizzeta Heart Container": "SnowpeakRuins",
"Snowpeak Ruins Broken Floor Chest": "SnowpeakRuins",
"Snowpeak Ruins Chapel Chest": "SnowpeakRuins",
"Snowpeak Ruins Chest After Darkhammer": "SnowpeakRuins",
"Snowpeak Ruins Courtyard Central Chest": "SnowpeakRuins",
"Snowpeak Ruins Dungeon Reward": "SnowpeakRuins",
"Snowpeak Ruins East Courtyard Buried Chest": "SnowpeakRuins",
"Snowpeak Ruins East Courtyard Chest": "SnowpeakRuins",
"Snowpeak Ruins Ice Room Poe": "SnowpeakRuins",
"Snowpeak Ruins Lobby Armor Poe": "SnowpeakRuins",
"Snowpeak Ruins Lobby Chandelier Chest": "SnowpeakRuins",
"Snowpeak Ruins Lobby East Armor Chest": "SnowpeakRuins",
"Snowpeak Ruins Lobby Poe": "SnowpeakRuins",
"Snowpeak Ruins Lobby West Armor Chest": "SnowpeakRuins",
"Snowpeak Ruins Mansion Map": "SnowpeakRuins",
"Snowpeak Ruins Northeast Chandelier Chest": "SnowpeakRuins",
"Snowpeak Ruins Ordon Pumpkin Chest": "SnowpeakRuins",
"Snowpeak Ruins West Cannon Room Central Chest": "SnowpeakRuins",
"Snowpeak Ruins West Cannon Room Corner Chest": "SnowpeakRuins",
"Snowpeak Ruins West Courtyard Buried Chest": "SnowpeakRuins",
"Snowpeak Ruins Wooden Beam Central Chest": "SnowpeakRuins",
"Snowpeak Ruins Wooden Beam Chandelier Chest": "SnowpeakRuins",
"Snowpeak Ruins Wooden Beam Northwest Chest": "SnowpeakRuins",
"Temple of Time Armogohma Heart Container": "TempleOfTime",
"Temple of Time Armos Antechamber East Chest": "TempleOfTime",
"Temple of Time Armos Antechamber North Chest": "TempleOfTime",
"Temple of Time Armos Antechamber Statue Chest": "TempleOfTime",
"Temple of Time Big Key Chest": "TempleOfTime",
"Temple of Time Chest Before Darknut": "TempleOfTime",
"Temple of Time Darknut Chest": "TempleOfTime",
"Temple of Time Dungeon Reward": "TempleOfTime",
"Temple of Time First Staircase Armos Chest": "TempleOfTime",
"Temple of Time First Staircase Gohma Gate Chest": "TempleOfTime",
"Temple of Time First Staircase Window Chest": "TempleOfTime",
"Temple of Time Floor Switch Puzzle Room Upper Chest": "TempleOfTime",
"Temple of Time Gilloutine Chest": "TempleOfTime",
"Temple of Time Lobby Lantern Chest": "TempleOfTime",
"Temple of Time Moving Wall Beamos Room Chest": "TempleOfTime",
"Temple of Time Moving Wall Dinalfos Room Chest": "TempleOfTime",
"Temple of Time Poe Above Scales": "TempleOfTime",
"Temple of Time Poe Behind Gate": "TempleOfTime",
"Temple of Time Scales Gohma Chest": "TempleOfTime",
"Temple of Time Scales Upper Chest": "TempleOfTime",

}

# Exclude required dungeons checks from all_dungeon_checks
dungeon_checks_barren = set()
for check, zone in all_dungeon_checks.items():
    if zone in dungeons:
        dungeon_checks_barren.add(check)

# List of checks to exclude for the Sometimes category        
excluded_checks = ["Auru Gift To Fyer",
"Cats Hide and Seek Minigame",
"Ilia Charm",
"Ilia Memory Reward",
"Iza Raging Rapids Minigame",
"Jovani 60 Poe Soul Reward",
"Ordon Cat Rescue",
"Ordon Shield",
"Ordon Sword",
"Renados Letter",
"Rutelas Blessing",
"Skybook From Impaz",
"Telma Invoice",
"Uli Cradle Delivery",
"Wooden Statue",
"West Hyrule Field Golden Wolf",
"Faron Woods Golden Wolf",
"Gerudo Desert Golden Wolf",
"Kakariko Graveyard Golden Wolf",
"North Castle Town Golden Wolf",
"Ordon Spring Golden Wolf",
"Outside South Castle Town Golden Wolf",
"Agitha Female Ant Reward",
"Agitha Female Beetle Reward",
"Agitha Female Butterfly Reward",
"Agitha Female Dayfly Reward",
"Agitha Female Dragonfly Reward",
"Agitha Female Grasshopper Reward",
"Agitha Female Ladybug Reward",
"Agitha Female Mantis Reward",
"Agitha Female Phasmid Reward",
"Agitha Female Pill Bug Reward",
"Agitha Female Snail Reward",
"Agitha Female Stag Beetle Reward",
"Agitha Male Ant Reward",
"Agitha Male Beetle Reward",
"Agitha Male Butterfly Reward",
"Agitha Male Dayfly Reward",
"Agitha Male Dragonfly Reward",
"Agitha Male Grasshopper Reward",
"Agitha Male Ladybug Reward",
"Agitha Male Mantis Reward",
"Agitha Male Phasmid Reward",
"Agitha Male Pill Bug Reward",
"Agitha Male Snail Reward",
"Agitha Male Stag Beetle Reward",
"Cave of Ordeals Floor 17 Poe",
"Cave of Ordeals Floor 33 Poe",
"Cave of Ordeals Floor 44 Poe",
"Cave of Ordeals Great Fairy Reward",
"Hidden Village Poe",
"Doctors Office Balcony Chest",
"Snowboard Racing Prize"
"Progressive_Fused_Shadow",
"Progressive_Mirror_Shard",
"Arbiters Grounds East Turning Room Poe",
"Arbiters Grounds Hidden Wall Poe",
"Arbiters Grounds Torch Room Poe",
"Arbiters Grounds West Poe",
"Bulblin Camp Poe",
"City in The Sky Garden Island Poe",
"City in The Sky Poe Above Central Fan",
"Death Mountain Trail Poe",
"East Castle Town Bridge Poe",
"Eldin Lantern Cave Poe",
"Faron Field Poe",
"Faron Mist Poe",
"Flight By Fowl Ledge Poe",
"Gerudo Desert East Poe",
"Gerudo Desert North Peahat Poe",
"Gerudo Desert Poe Above Cave of Ordeals",
"Gerudo Desert Rock Grotto First Poe",
"Gerudo Desert Rock Grotto Second Poe",
"Hidden Village Poe",
"Hyrule Field Ampitheater Poe",
"Isle of Riches Poe",
"Jovani House Poe",
"Kakariko Gorge Poe",
"Kakariko Graveyard Grave Poe",
"Kakariko Graveyard Open Poe",
"Kakariko Village Bomb Shop Poe",
"Kakariko Village Watchtower Poe",
"Lake Hylia Alcove Poe",
"Lake Hylia Bridge Cliff Poe",
"Lake Hylia Dock Poe",
"Lake Hylia Tower Poe",
"Lake Lantern Cave Final Poe",
"Lake Lantern Cave First Poe",
"Lake Lantern Cave Second Poe",
"Lanayru Field Bridge Poe",
"Lanayru Field Poe Grotto Left Poe",
"Lanayru Field Poe Grotto Right Poe",
"Lost Woods Boulder Poe",
"Lost Woods Waterfall Poe",
"Outside Arbiters Grounds Poe",
"Outside Bulblin Camp Poe",
"Outside South Castle Town Poe",
"Sacred Grove Master Sword Poe",
"Sacred Grove Temple of Time Owl Statue Poe",
"Snowpeak Above Freezard Grotto Poe",
"Snowpeak Blizzard Poe",
"Snowpeak Cave Ice Poe",
"Snowpeak Icy Summit Poe",
"Snowpeak Poe Among Trees",
"Snowpeak Ruins Ice Room Poe",
"Snowpeak Ruins Lobby Armor Poe",
"Snowpeak Ruins Lobby Poe",
"Temple of Time Poe Above Scales",
"Temple of Time Poe Behind Gate",
"Upper Zoras River Poe",
"Zoras Domain Mother and Child Isle Poe",
"Zoras Domain Waterfall Poe"]

# Create the list of barren_checks
barren_checks = [check for check, zone in zones.items() if zone in barren]

# Add to the excluded_checks list the barren dungeons checks
excluded_checks.extend(dungeon_checks_barren)
excluded_checks.extend(barren_checks)

# Select 3 Sometimes checks from "itemPlacements" but exclude the ones from excluded_checks
valid_items = {k: v for k, v in data["itemPlacements"].items() if k not in excluded_checks}
sometimes_items = dict(random.sample(list(valid_items.items()), 3))

# Prompt Sometimes
for check, item in sometimes_items.items():
    print(f"{check} : {item}")
    with open('Hint_list.txt', "a") as file:
        file.write(f"{check} : {item}\n")

# Add Sometimes to the xlsx file
for i, (check, item) in enumerate(sometimes_items.items(), start=18):
    worksheet.cell(row=i, column=2, value=f'{check} : {item}')

# Save the file
workbook.save('Hint_list.xlsx')
